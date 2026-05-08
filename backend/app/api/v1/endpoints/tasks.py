# ## Imports ─────────────────────────────────────────────────────────────────
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone
import uuid
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.db.database import get_db
from app.models.task import Task, TaskMaterial, TaskQuestionBank, TaskPersona, TaskAssignment, TaskSubmission, MaterialReadingProgress, RoleplaySession, SubmissionAnswer
from app.models.file import File
from app.models.notification import Notification
from app.models.user import User
from app.models.question import Question
from app.models.persona import Persona
from app.core.deps import get_current_user, require_admin
from app.core.security import decode_token

router = APIRouter(prefix="/tasks", tags=["任务"])


# ## 请求 Schema ───────────────────────────────────────────────────────────
class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    type: str  # qa / roleplay
    config: dict = {}
    start_at: Optional[str] = None
    end_at: Optional[str] = None
    material_ids: list[str] = []
    bank_ids: list[str] = []
    persona_ids: list[str] = []


class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    config: Optional[dict] = None
    start_at: Optional[str] = None
    end_at: Optional[str] = None


# ## 序列化工具 ────────────────────────────────────────────────────────────
def task_to_dict(t: Task) -> dict:
    return {
        "id": t.id, "title": t.title, "description": t.description,
        "type": t.type, "status": t.status, "config": t.config,
        "start_at": t.start_at.isoformat() if t.start_at else None,
        "end_at": t.end_at.isoformat() if t.end_at else None,
        "created_at": t.created_at.isoformat(),
        "material_ids": [m.file_id for m in t.materials],
    }


# ## 管理员：任务列表 ─────────────────────────────────────────────────────
@router.get("")
async def list_tasks(
    type: Optional[str] = None, status: Optional[str] = None,
    page: int = 1, page_size: int = 20,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
):
    q = select(Task)
    if type:
        q = q.where(Task.type == type)
    if status:
        q = q.where(Task.status == status)
    total = await db.scalar(select(func.count()).select_from(q.subquery()))
    result = await db.execute(q.order_by(Task.created_at.desc()).offset((page - 1) * page_size).limit(page_size))
    tasks = result.scalars().all()

    task_ids = [t.id for t in tasks]
    count_rows = await db.execute(
        select(TaskAssignment.task_id, func.count(TaskAssignment.id).label("cnt"))
        .where(TaskAssignment.task_id.in_(task_ids))
        .group_by(TaskAssignment.task_id)
    )
    count_map = {row.task_id: row.cnt for row in count_rows}

    items = []
    for t in tasks:
        d = task_to_dict(t)
        d["assigned_count"] = count_map.get(t.id, 0)
        items.append(d)
    return {"total": total, "items": items}


# ## 创建任务 ──────────────────────────────────────────────────────────────
@router.post("", status_code=201)
async def create_task(data: TaskCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    task = Task(
        id=str(uuid.uuid4()), title=data.title, description=data.description,
        type=data.type, config=data.config, created_by=current_user.id,
        start_at=datetime.fromisoformat(data.start_at) if data.start_at else None,
        end_at=datetime.fromisoformat(data.end_at) if data.end_at else None,
    )
    db.add(task)
    await db.flush()

    for i, fid in enumerate(data.material_ids):
        db.add(TaskMaterial(id=str(uuid.uuid4()), task_id=task.id, file_id=fid, sort_order=i))
    for bid in data.bank_ids:
        db.add(TaskQuestionBank(task_id=task.id, bank_id=bid))
    for pid in data.persona_ids:
        db.add(TaskPersona(task_id=task.id, persona_id=pid))

    await db.commit()
    await db.refresh(task)
    return task_to_dict(task)


# ## 学员：我的任务列表 ────────────────────────────────────────────────────
@router.get("/my")
async def my_tasks(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    assign_result = await db.execute(
        select(TaskAssignment, Task)
        .join(Task, TaskAssignment.task_id == Task.id)
        .where(TaskAssignment.user_id == current_user.id, Task.status == "active")
        .order_by(Task.created_at.desc())
    )
    rows = assign_result.all()

    task_ids = [task.id for _, task in rows]
    sub_result = await db.execute(
        select(TaskSubmission)
        .where(TaskSubmission.task_id.in_(task_ids), TaskSubmission.user_id == current_user.id)
        .order_by(TaskSubmission.started_at.desc())
    )
    sub_map: dict = {}
    for s in sub_result.scalars().all():
        if s.task_id not in sub_map:
            sub_map[s.task_id] = s

    items = []
    for assignment, task in rows:
        sub = sub_map.get(task.id)
        if sub and sub.status == "completed":
            assign_status = "completed"
        elif sub:
            assign_status = "started"
        else:
            assign_status = "assigned"
        d = task_to_dict(task)
        items.append({
            "id": assignment.id, "task_id": task.id,
            "user_id": assignment.user_id, "status": assign_status, "task": d,
        })
    return items


# ## 任务详情 ──────────────────────────────────────────────────────────────
@router.get("/{task_id}")
async def get_task(task_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "任务不存在")

    banks = await db.execute(select(TaskQuestionBank).where(TaskQuestionBank.task_id == task_id))
    personas = await db.execute(select(TaskPersona).where(TaskPersona.task_id == task_id))

    d = task_to_dict(task)
    d["bank_ids"] = [r.bank_id for r in banks.scalars().all()]
    d["persona_ids"] = [r.persona_id for r in personas.scalars().all()]
    d["assigned_user_ids"] = [a.user_id for a in task.assignments]
    return d


# ## 更新任务 ──────────────────────────────────────────────────────────────
@router.put("/{task_id}")
async def update_task(task_id: str, data: TaskUpdate, db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "任务不存在")
    for field, val in data.model_dump(exclude_none=True).items():
        if field in ("start_at", "end_at") and val:
            setattr(task, field, datetime.fromisoformat(val))
        else:
            setattr(task, field, val)
    await db.commit()
    return task_to_dict(task)


# ## 发布任务 & 指派学员 ────────────────────────────────────────────────────
@router.post("/{task_id}/publish")
async def publish_task(task_id: str, body: dict, db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "任务不存在")

    user_ids = body.get("user_ids", [])
    task.status = "active"

    existing_res = await db.execute(select(TaskAssignment.user_id).where(TaskAssignment.task_id == task_id))
    existing_ids = set(existing_res.scalars().all())

    new_count = 0
    for uid in user_ids:
        if uid not in existing_ids:
            db.add(TaskAssignment(id=str(uuid.uuid4()), task_id=task_id, user_id=uid))
            db.add(Notification(
                id=str(uuid.uuid4()), user_id=uid, type="task_assigned",
                title=f"新任务：{task.title}", body="您有一个新的培训任务，请及时完成。",
                task_id=task_id,
            ))
            new_count += 1
    await db.commit()
    return {"ok": True, "assigned": new_count, "total": len(user_ids)}


# ## 任务完成进度（汇总）────────────────────────────────────────────────────
@router.get("/{task_id}/progress")
async def task_progress(task_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    assignments = await db.execute(select(TaskAssignment).where(TaskAssignment.task_id == task_id))
    assigned_ids = [a.user_id for a in assignments.scalars().all()]

    completed = await db.scalar(
        select(func.count()).select_from(TaskSubmission)
        .where(TaskSubmission.task_id == task_id, TaskSubmission.status == "completed")
    )
    return {
        "total_assigned": len(assigned_ids),
        "completed": completed,
        "completion_rate": round(completed / len(assigned_ids) * 100, 1) if assigned_ids else 0,
    }


# ## 任务完成进度（逐人明细）──────────────────────────────────────────────
@router.get("/{task_id}/progress-detail")
async def task_progress_detail(task_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "任务不存在")

    assign_res = await db.execute(
        select(TaskAssignment, User)
        .join(User, TaskAssignment.user_id == User.id)
        .where(TaskAssignment.task_id == task_id)
    )
    rows = assign_res.all()

    sub_res = await db.execute(
        select(TaskSubmission)
        .where(TaskSubmission.task_id == task_id)
        .order_by(TaskSubmission.started_at.desc())
    )
    sub_map: dict = {}
    for s in sub_res.scalars().all():
        if s.user_id not in sub_map:
            sub_map[s.user_id] = s

    now = datetime.now(timezone.utc)
    items = []
    for assignment, user in rows:
        sub = sub_map.get(user.id)
        if sub and sub.status == "completed":
            status = "completed"
        elif sub:
            status = "in_progress"
        elif task.end_at and task.end_at.replace(tzinfo=timezone.utc if task.end_at.tzinfo is None else task.end_at.tzinfo) < now:
            status = "expired"
        else:
            status = "not_started"

        items.append({
            "user_id": user.id,
            "user": {"username": user.username, "display_name": user.display_name},
            "status": status,
            "score": float(sub.score) if sub and sub.score is not None else None,
            "max_score": float(sub.max_score) if sub and sub.max_score is not None else None,
            "started_at": sub.started_at.isoformat() if sub else None,
            "completed_at": sub.completed_at.isoformat() if sub and sub.completed_at else None,
        })
    return {"total": len(items), "items": items}


# ## 能力雷达：维度得分汇总 ────────────────────────────────────────────────
@router.get("/{task_id}/dimension-stats")
async def task_dimension_stats(task_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    """Aggregate dimension scores from final_evaluation of completed roleplay sessions."""
    # Get all TaskSubmissions for this task
    sub_res = await db.execute(
        select(TaskSubmission).where(
            TaskSubmission.task_id == task_id,
            TaskSubmission.status == "completed"
        )
    )
    submissions = sub_res.scalars().all()

    # Collect dimension scores from RoleplaySessions
    dim_totals: dict[str, list[float]] = {}
    dim_max: dict[str, int] = {}

    for sub in submissions:
        # Get roleplay session for this submission
        session_res = await db.execute(
            select(RoleplaySession).where(RoleplaySession.submission_id == sub.id)
        )
        session = session_res.scalar_one_or_none()
        if not session or not session.final_evaluation:
            continue

        dims = session.final_evaluation.get("dimensions", [])
        for d in dims:
            name = d.get("name", "")
            score = d.get("score")
            max_val = d.get("max", 100)
            if name and score is not None:
                if name not in dim_totals:
                    dim_totals[name] = []
                    dim_max[name] = max_val
                dim_totals[name].append(float(score))

    if not dim_totals:
        return {"dimensions": [], "session_count": 0}

    dimensions = [
        {
            "name": name,
            "avg_score": round(sum(scores) / len(scores), 1),
            "max": dim_max[name],
            "count": len(scores),
        }
        for name, scores in dim_totals.items()
    ]

    return {"dimensions": dimensions, "session_count": len(submissions)}


# ## 任务关联资源更新（素材/题库/角色）────────────────────────────────────
@router.put("/{task_id}/associations")
async def update_task_associations(task_id: str, body: dict, db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "任务不存在")

    material_ids = body.get("material_ids")
    bank_ids = body.get("bank_ids")
    persona_ids = body.get("persona_ids")

    if material_ids is not None:
        await db.execute(delete(TaskMaterial).where(TaskMaterial.task_id == task_id))
        for i, fid in enumerate(material_ids):
            db.add(TaskMaterial(id=str(uuid.uuid4()), task_id=task_id, file_id=fid, sort_order=i))

    if bank_ids is not None:
        await db.execute(delete(TaskQuestionBank).where(TaskQuestionBank.task_id == task_id))
        for bid in bank_ids:
            db.add(TaskQuestionBank(task_id=task_id, bank_id=bid))

    if persona_ids is not None:
        await db.execute(delete(TaskPersona).where(TaskPersona.task_id == task_id))
        for pid in persona_ids:
            db.add(TaskPersona(task_id=task_id, persona_id=pid))

    await db.commit()
    banks_res = await db.execute(select(TaskQuestionBank).where(TaskQuestionBank.task_id == task_id))
    personas_res = await db.execute(select(TaskPersona).where(TaskPersona.task_id == task_id))
    d = task_to_dict(task)
    d["bank_ids"] = [r.bank_id for r in banks_res.scalars().all()]
    d["persona_ids"] = [r.persona_id for r in personas_res.scalars().all()]
    return d


# ## 素材阅读进度上报 ─────────────────────────────────────────────────────
@router.post("/{task_id}/materials/{file_id}/page")
async def report_reading_page(
    task_id: str, file_id: str, body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    page = body.get("page", 1)
    total_pages = body.get("total_pages", 1)

    result = await db.execute(
        select(MaterialReadingProgress).where(
            MaterialReadingProgress.user_id == current_user.id,
            MaterialReadingProgress.task_id == task_id,
            MaterialReadingProgress.file_id == file_id,
        )
    )
    progress = result.scalar_one_or_none()

    if not progress:
        progress = MaterialReadingProgress(
            id=str(uuid.uuid4()), user_id=current_user.id,
            task_id=task_id, file_id=file_id, total_pages=total_pages,
        )
        db.add(progress)

    visited = set(progress.visited_pages or [])
    visited.add(page)
    progress.visited_pages = list(visited)
    progress.max_page_reached = max(progress.max_page_reached, page)
    progress.total_pages = total_pages
    progress.is_completed = len(visited) >= total_pages
    progress.last_read_at = datetime.now(timezone.utc)
    await db.commit()
    return {"visited": len(visited), "total": total_pages, "is_completed": progress.is_completed}


# ## 素材阅读状态查询 ─────────────────────────────────────────────────────
@router.get("/{task_id}/reading-status")
async def reading_status(task_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(
        select(MaterialReadingProgress).where(
            MaterialReadingProgress.user_id == current_user.id,
            MaterialReadingProgress.task_id == task_id,
        )
    )
    items = result.scalars().all()
    return {p.file_id: {"is_completed": p.is_completed, "visited": len(p.visited_pages or [])} for p in items}


# ## 任务结果导出（Excel）────────────────────────────────────────────────────
# 支持 token query 参数，前端直接用 <a href> 下载时无法带 Authorization 头
@router.get("/{task_id}/export")
async def export_task_results(
    task_id: str,
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    # token 验证（管理员权限）
    if not token:
        raise HTTPException(401, "缺少 token")
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "无效 token")
    user_id = payload.get("sub")
    user = await db.get(User, user_id)
    if not user or user.role not in ("admin", "superadmin"):
        raise HTTPException(403, "无权限")

    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(404, "任务不存在")

    # 查询所有指派用户
    assign_res = await db.execute(
        select(TaskAssignment, User)
        .join(User, TaskAssignment.user_id == User.id)
        .where(TaskAssignment.task_id == task_id)
    )
    assign_rows = assign_res.all()

    # 查询所有提交记录
    sub_res = await db.execute(
        select(TaskSubmission)
        .where(TaskSubmission.task_id == task_id)
        .order_by(TaskSubmission.started_at.asc())
    )
    submissions = sub_res.scalars().all()

    # 按 user_id 分组，practice / exam 各取最新一条
    practice_map: dict[str, TaskSubmission] = {}
    exam_map: dict[str, TaskSubmission] = {}
    for s in submissions:
        if s.mode == "practice" and s.status == "completed":
            practice_map[s.user_id] = s
        elif s.mode in ("exam", "assessment") and s.status == "completed":
            exam_map[s.user_id] = s

    wb = openpyxl.Workbook()

    # ── 样式定义 ──────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header_row(ws, col_count: int):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        ws.row_dimensions[1].height = 22

    def fmt_dt(dt) -> str:
        if not dt:
            return "—"
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M")

    def fmt_score(score, max_score) -> str:
        if score is None:
            return "—"
        s = f"{float(score):.1f}"
        if max_score:
            s += f" / {float(max_score):.0f}"
        return s

    def fmt_duration(started_at, completed_at) -> str:
        if not started_at or not completed_at:
            return "—"
        if started_at.tzinfo is None:
            started_at = started_at.replace(tzinfo=timezone.utc)
        if completed_at.tzinfo is None:
            completed_at = completed_at.replace(tzinfo=timezone.utc)
        secs = int((completed_at - started_at).total_seconds())
        if secs < 60:
            return f"{secs}秒"
        return f"{secs // 60}分{secs % 60}秒"

    is_qa = task.type == "qa"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1：学员总览
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "学员总览"

    if is_qa:
        headers1 = ["姓名", "账号", "练习得分", "练习完成时间", "练习用时", "考核得分", "考核完成时间", "考核用时", "整体状态"]
        col_widths1 = [16, 16, 12, 18, 10, 12, 18, 10, 10]
    else:
        headers1 = ["姓名", "账号", "练习得分", "练习完成时间", "练习轮次", "考核得分", "考核完成时间", "考核轮次", "整体状态"]
        col_widths1 = [16, 16, 12, 18, 10, 12, 18, 10, 10]

    ws1.append(headers1)
    style_header_row(ws1, len(headers1))
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[ws1.cell(1, i).column_letter].width = w

    # 收集 roleplay session 用于轮次数
    session_map: dict[str, RoleplaySession] = {}
    if not is_qa:
        all_sub_ids = [s.id for s in submissions]
        if all_sub_ids:
            sess_res = await db.execute(
                select(RoleplaySession).where(RoleplaySession.submission_id.in_(all_sub_ids))
            )
            for sess in sess_res.scalars().all():
                session_map[sess.submission_id] = sess

    for _, u in assign_rows:
        pr = practice_map.get(u.id)
        ex = exam_map.get(u.id)
        name = u.display_name or u.username or u.phone or "—"
        account = u.username or u.phone or "—"

        if is_qa:
            pr_score = fmt_score(pr.score, pr.max_score) if pr else "—"
            pr_time = fmt_dt(pr.completed_at) if pr else "—"
            pr_dur = fmt_duration(pr.started_at, pr.completed_at) if pr else "—"
            ex_score = fmt_score(ex.score, ex.max_score) if ex else "—"
            ex_time = fmt_dt(ex.completed_at) if ex else "—"
            ex_dur = fmt_duration(ex.started_at, ex.completed_at) if ex else "—"
        else:
            pr_sess = session_map.get(pr.id) if pr else None
            ex_sess = session_map.get(ex.id) if ex else None
            pr_score = fmt_score(pr.score, pr.max_score) if pr else "—"
            pr_time = fmt_dt(pr.completed_at) if pr else "—"
            pr_dur = str(pr_sess.total_turns) + " 轮" if pr_sess else "—"
            ex_score = fmt_score(ex.score, ex.max_score) if ex else "—"
            ex_time = fmt_dt(ex.completed_at) if ex else "—"
            ex_dur = str(ex_sess.total_turns) + " 轮" if ex_sess else "—"

        if ex:
            overall = "考核已完成"
        elif pr:
            overall = "练习已完成"
        else:
            overall = "未开始"

        ws1.append([name, account, pr_score, pr_time, pr_dur, ex_score, ex_time, ex_dur, overall])

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=len(headers1)):
        for cell in row:
            cell.alignment = left

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2：QA 题目明细 / Roleplay 维度评分
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("明细")

    if is_qa:
        # 收集所有考核提交的答案
        exam_sub_ids = [s.id for s in exam_map.values()]
        practice_sub_ids = [s.id for s in practice_map.values()]
        all_detail_ids = list(set(exam_sub_ids + practice_sub_ids))

        answers_res = await db.execute(
            select(SubmissionAnswer)
            .where(SubmissionAnswer.submission_id.in_(all_detail_ids))
            .order_by(SubmissionAnswer.submission_id, SubmissionAnswer.question_order)
        ) if all_detail_ids else None
        answers = answers_res.scalars().all() if answers_res else []

        # 查题目文本
        q_ids = list({a.question_id for a in answers if a.question_id})
        q_map: dict[str, Question] = {}
        if q_ids:
            q_res = await db.execute(select(Question).where(Question.id.in_(q_ids)))
            for q in q_res.scalars().all():
                q_map[q.id] = q

        # sub_id → (user, mode)
        sub_user_map: dict[str, tuple] = {}
        for _, u in assign_rows:
            if u.id in practice_map:
                sub_user_map[practice_map[u.id].id] = (u, "练习")
            if u.id in exam_map:
                sub_user_map[exam_map[u.id].id] = (u, "考核")

        headers2 = ["姓名", "账号", "模式", "题号", "题目", "作答（语音转写）", "得分", "AI 评语", "参考答案"]
        col_widths2 = [14, 14, 8, 6, 30, 30, 10, 30, 30]
        ws2.append(headers2)
        style_header_row(ws2, len(headers2))
        for i, w in enumerate(col_widths2, 1):
            ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

        for ans in answers:
            user_info = sub_user_map.get(ans.submission_id)
            if not user_info:
                continue
            u, mode = user_info
            q = q_map.get(ans.question_id) if ans.question_id else None
            q_text = q.question_text if q else "—"
            name = u.display_name or u.username or u.phone or "—"
            account = u.username or u.phone or "—"
            ws2.append([
                name, account, mode,
                ans.question_order or "—",
                q_text,
                ans.transcribed_text or "—",
                fmt_score(ans.score, ans.max_score),
                ans.ai_feedback or "—",
                ans.reference_answer or "—",
            ])

        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(headers2)):
            for cell in row:
                cell.alignment = left

    else:
        # Roleplay 维度评分明细
        headers2 = ["姓名", "账号", "模式", "完成时间", "综合得分", "亮点", "改进建议", "维度1", "得分1", "维度2", "得分2", "维度3", "得分3", "维度4", "得分4"]
        col_widths2 = [14, 14, 8, 18, 10, 30, 30, 14, 8, 14, 8, 14, 8, 14, 8]
        ws2.append(headers2)
        style_header_row(ws2, len(headers2))
        for i, w in enumerate(col_widths2, 1):
            ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

        user_map_by_id = {u.id: u for _, u in assign_rows}

        def write_roleplay_row(sub: TaskSubmission, mode_label: str):
            u = user_map_by_id.get(sub.user_id)
            if not u:
                return
            sess = session_map.get(sub.id)
            name = u.display_name or u.username or u.phone or "—"
            account = u.username or u.phone or "—"
            if not sess or not sess.final_evaluation:
                ws2.append([name, account, mode_label, fmt_dt(sub.completed_at),
                             fmt_score(sub.score, sub.max_score), "—", "—"] + ["—"] * 8)
                return
            ev = sess.final_evaluation
            highlights = "；".join(ev.get("highlights", []))
            improvements = "；".join(ev.get("improvements", []))
            dims = ev.get("dimensions", [])[:4]
            dim_cols = []
            for d in dims:
                dim_cols += [d.get("name", "—"), f"{d.get('score', '—')} / {d.get('max', 100)}"]
            while len(dim_cols) < 8:
                dim_cols.append("—")
            ws2.append([name, account, mode_label, fmt_dt(sub.completed_at),
                         fmt_score(sub.score, sub.max_score), highlights, improvements] + dim_cols)

        for u_id, sub in practice_map.items():
            write_roleplay_row(sub, "练习")
        for u_id, sub in exam_map.items():
            write_roleplay_row(sub, "考核")

        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(headers2)):
            for cell in row:
                cell.alignment = left

    # ── 元信息 Sheet ─────────────────────────────────────────────────────
    ws_info = wb.create_sheet("任务信息")
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 40
    info_rows = [
        ("任务名称", task.title),
        ("任务类型", "知识问答" if is_qa else "模拟对练"),
        ("任务状态", task.status),
        ("截止时间", fmt_dt(task.end_at) if task.end_at else "无"),
        ("总指派人数", len(assign_rows)),
        ("练习完成人数", len(practice_map)),
        ("考核完成人数", len(exam_map)),
        ("导出时间", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for k, v in info_rows:
        ws_info.append([k, str(v)])
    for row in ws_info.iter_rows(min_row=1, max_row=len(info_rows), min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[0].alignment = left
        row[1].alignment = left

    # ── 输出 ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = task.title.replace("/", "_").replace("\\", "_")[:40]
    filename = f"任务结果_{safe_title}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
